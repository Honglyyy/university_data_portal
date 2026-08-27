from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Count, Model

from myapp.models import Teacher, Class, Student


@login_required(login_url='/login')
def index(request):
    user = request.user
    if hasattr(user, 'teacher'):
        return redirect('teacher_dashboard')
    elif hasattr(user, 'student'):
        return redirect('student_dashboard')
    elif user.is_superuser:
        return redirect('/admin/')
    elif user.is_staff:
        return redirect('frontdesk_dashboard')
    else:
        # Fallback if the user has no role assigned
        return render(request, 'myapp/base.html')

@login_required(login_url='/login')
def teacher_dashboard(request):
    try:
        teacher = request.user.teacher
        classes = teacher.class_set.select_related('subject', 'term', 'room').annotate(
            student_count=Count('enrollment')
        ).order_by('-term__start_date', 'subject__name')
        salary = teacher.salary
    except AttributeError:
        classes = []
        salary = None

    context = {
        'classes': classes,
        'salary': salary,
    }

    return render(request, 'myapp/teacher_dashboard.html', context)

@login_required(login_url='/login')
def teacher_classes(request):
    try:
        teacher = request.user.teacher
        classes = teacher.class_set.select_related('subject', 'term', 'room').annotate(
            student_count=Count('enrollment')
        ).order_by('-term__start_date', 'subject__name')
        
        grouped_classes = {}
        for c in classes:
            if c.term not in grouped_classes:
                grouped_classes[c.term] = []
            grouped_classes[c.term].append(c)
    except AttributeError:
        classes = []
        grouped_classes = {}

    context = {
        'classes': classes,
        'grouped_classes': grouped_classes,
    }

    return render(request, 'myapp/teacher_classes.html', context)

@login_required(login_url='/login')
def student_dashboard(request):
    try:
        # Get the enrollments for the logged-in student
        enrollments = request.user.student.enrollment_set.select_related(
            'student',
            'class_instance__subject',
            'class_instance__term',
            'class_instance__room',
            'class_instance__teacher',
        ).order_by('-class_instance__term__start_date', '-class_instance__term__name', 'class_instance__subject__name')
        
        # Calculate total credits
        total_credits = sum(e.class_instance.subject.credits for e in enrollments)
        
        # Get current term from the first enrollment if available
        first_enrollment = enrollments.first()
        current_term = first_enrollment.class_instance.term if first_enrollment else None
    except AttributeError:
        enrollments = []
        total_credits = 0
        current_term = None

    context = {
        'enrollments': enrollments,
        'total_credits': total_credits,
        'current_term': current_term,
    }
    return render(request, 'myapp/student_dashboard.html', context)

@login_required(login_url='/login')
def student_major(request):
    try:
        student = request.user.student
    except AttributeError:
        student = None
    
    context = {
        'student': student
    }
    return render(request, 'myapp/student_major.html', context)

@login_required(login_url='/login')
def student_grades(request):
    try:
        enrollments = request.user.student.enrollment_set.select_related(
            'class_instance__subject',
            'class_instance__term'
        ).order_by('-class_instance__term__start_date', '-class_instance__term__name', 'class_instance__subject__name')
        
        overall_total_credits = 0
        overall_credits_earned = 0
        overall_total_points = 0
        overall_graded_credits = 0
        
        # We will use an ordered dictionary to keep the sorting order
        terms_data_dict = {}
        
        for e in enrollments:
            term = e.class_instance.term
            if term not in terms_data_dict:
                terms_data_dict[term] = {
                    'term': term,
                    'enrollments': [],
                    'total_credits': 0,
                    'credits_earned': 0,
                    'total_points': 0,
                    'graded_credits': 0,
                }
            
            terms_data_dict[term]['enrollments'].append(e)
            
            subject_credits = e.class_instance.subject.credits
            terms_data_dict[term]['total_credits'] += subject_credits
            overall_total_credits += subject_credits
            
            if e.final_grade:
                e.point = e.gpa
                e.total_point = e.gpa * subject_credits
                
                terms_data_dict[term]['total_points'] += e.total_point
                overall_total_points += e.total_point
                
                terms_data_dict[term]['graded_credits'] += subject_credits
                overall_graded_credits += subject_credits
                
                terms_data_dict[term]['credits_earned'] += e.credits
                overall_credits_earned += e.credits
            else:
                e.point = None
                e.total_point = None
                
        terms_data_list = []
        for term, data in terms_data_dict.items():
            g_credits = data['graded_credits']
            data['gpa'] = data['total_points'] / g_credits if g_credits > 0 else 0.0
            terms_data_list.append(data)
            
        overall_gpa = overall_total_points / overall_graded_credits if overall_graded_credits > 0 else 0.0
        
    except AttributeError:
        terms_data_list = []
        overall_total_credits = 0
        overall_credits_earned = 0
        overall_gpa = 0.0
        
    context = {
        'terms_data': terms_data_list,
        'overall_total_credits': overall_total_credits,
        'overall_credits_earned': overall_credits_earned,
        'overall_gpa': round(overall_gpa, 2)
    }
    return render(request, 'myapp/student_grades.html', context)


@login_required(login_url='/login')
def teacher_roster(request, class_id):
    try:
        teacher = request.user.teacher
        # Ensure the class belongs to this teacher
        class_instance = Class.objects.get(id=class_id, teacher=teacher)
    except (AttributeError, Class.DoesNotExist):
        return redirect('teacher_dashboard')

    expected_batch = class_instance.computed_batch
    enrollments = class_instance.enrollment_set.filter(
        student__batch=expected_batch
    ).select_related('student', 'student__user').all()

    grade_distribution = {'A+': 0, 'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'N/A': 0}
    for e in enrollments:
        if e.final_grade in grade_distribution:
            grade_distribution[e.final_grade] += 1
        else:
            grade_distribution['N/A'] += 1
            
    import json
    
    context = {
        'class_instance': class_instance,
        'enrollments': enrollments,
        'grades_labels_json': json.dumps(list(grade_distribution.keys())),
        'grades_counts_json': json.dumps(list(grade_distribution.values())),
    }
    return render(request, 'myapp/teacher_roster.html', context)

from django.views.decorators.http import require_POST
from myapp.models import Assessment, StudentScore

@login_required(login_url='/login')
def teacher_assignments(request, class_id):
    try:
        teacher = request.user.teacher
        class_instance = Class.objects.get(id=class_id, teacher=teacher)
    except (AttributeError, Class.DoesNotExist):
        return redirect('teacher_dashboard')

    if request.method == 'POST':
        name = request.POST.get('name')
        max_score = request.POST.get('max_score')
        if name and max_score:
            try:
                max_score_float = float(max_score)
                if max_score_float > 0 and max_score_float <= 99999.99:
                    Assessment.objects.create(
                        class_instance=class_instance,
                        name=name,
                        max_score=max_score_float
                    )
                else:
                    messages.error(request, 'Max score must be between 0.01 and 99999.99.')
            except ValueError:
                messages.error(request, 'Invalid max score value.')
            return redirect('teacher_assignments', class_id=class_id)
            
    assignments = Assessment.objects.filter(class_instance=class_instance).order_by('id')
    
    context = {
        'class_instance': class_instance,
        'assignments': assignments,
    }
    return render(request, 'myapp/teacher_assignments.html', context)

@login_required(login_url='/login')
def teacher_assignment_scores(request, class_id, assessment_id):
    try:
        teacher = request.user.teacher
        class_instance = Class.objects.get(id=class_id, teacher=teacher)
        assessment = Assessment.objects.get(id=assessment_id, class_instance=class_instance)
    except (AttributeError, Class.DoesNotExist, Assessment.DoesNotExist):
        return redirect('teacher_dashboard')
        
    expected_batch = class_instance.computed_batch
    enrollments = class_instance.enrollment_set.filter(
        student__batch=expected_batch
    ).select_related('student', 'student__user').all()

    if request.method == 'POST':
        for enrollment in enrollments:
            score_val = request.POST.get(f'score_{enrollment.id}')
            if score_val:
                # Update or create
                StudentScore.objects.update_or_create(
                    enrollment=enrollment,
                    assessment=assessment,
                    defaults={'score': score_val}
                )
            elif score_val == "":
                # If explicitly cleared, delete the score manually to trigger delete() method
                for score in StudentScore.objects.filter(
                    enrollment=enrollment,
                    assessment=assessment
                ):
                    score.delete()
        return redirect('teacher_assignment_scores', class_id=class_id, assessment_id=assessment_id)

    # Pre-fetch scores for this assessment
    student_scores = StudentScore.objects.filter(
        assessment=assessment, 
        enrollment__in=enrollments
    )
    score_map = {s.enrollment_id: s.score for s in student_scores}
    
    # attach score to enrollment for the template
    for e in enrollments:
        e.current_score = score_map.get(e.id, '')

    context = {
        'class_instance': class_instance,
        'assessment': assessment,
        'enrollments': enrollments,
    }
    return render(request, 'myapp/teacher_assignment_scores.html', context)


@login_required(login_url='/login')
@require_POST
def delete_assessment(request, class_id, assessment_id):
    try:
        teacher = request.user.teacher
        class_instance = Class.objects.get(id=class_id, teacher=teacher)
        assessment = Assessment.objects.get(id=assessment_id, class_instance=class_instance)
        # Manually delete scores to trigger the custom delete() method
        for score in StudentScore.objects.filter(assessment=assessment):
            score.delete()
        assessment.delete()
    except (AttributeError, Class.DoesNotExist, Assessment.DoesNotExist):
        pass
    
    return redirect('teacher_assignments', class_id=class_id)


import openpyxl
from django.http import HttpResponse
from django.contrib import messages

@login_required(login_url='/login')
def export_scores_excel(request, class_id, assessment_id):
    try:
        teacher = request.user.teacher
        class_instance = Class.objects.get(id=class_id, teacher=teacher)
        assessment = Assessment.objects.get(id=assessment_id, class_instance=class_instance)
    except (AttributeError, Class.DoesNotExist, Assessment.DoesNotExist):
        return redirect('teacher_dashboard')

    expected_batch = class_instance.computed_batch
    enrollments = class_instance.enrollment_set.filter(
        student__batch=expected_batch
    ).select_related('student', 'student__user').all()

    student_scores = StudentScore.objects.filter(
        assessment=assessment, 
        enrollment__in=enrollments
    )
    score_map = {s.enrollment_id: s.score for s in student_scores}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # Header
    ws.append(['Student ID', 'Student Name', 'Score'])

    for e in enrollments:
        student_id = e.student.student_id
        student_name = e.student.user.get_full_name() or e.student.user.username
        score = score_map.get(e.id, 0)
        ws.append([student_id, student_name, score])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{class_instance.subject.name}_{assessment.name}_Scores.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url='/login')
@require_POST
def import_scores_excel(request, class_id, assessment_id):
    try:
        teacher = request.user.teacher
        class_instance = Class.objects.get(id=class_id, teacher=teacher)
        assessment = Assessment.objects.get(id=assessment_id, class_instance=class_instance)
    except (AttributeError, Class.DoesNotExist, Assessment.DoesNotExist):
        return redirect('teacher_dashboard')

    if 'excel_file' not in request.FILES:
        messages.error(request, 'Please upload a valid Excel file.')
        return redirect('teacher_assignment_scores', class_id=class_id, assessment_id=assessment_id)

    excel_file = request.FILES['excel_file']
    if not excel_file.name.endswith('.xlsx'):
        messages.error(request, 'Invalid file format. Please upload an .xlsx file.')
        return redirect('teacher_assignment_scores', class_id=class_id, assessment_id=assessment_id)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        # Find indexes of columns by reading header
        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
        try:
            id_idx = headers.index('student id')
            score_idx = headers.index('score')
        except ValueError:
            messages.error(request, 'Invalid Excel format. Must contain "Student ID" and "Score" columns.')
            return redirect('teacher_assignment_scores', class_id=class_id, assessment_id=assessment_id)

        expected_batch = class_instance.computed_batch
        enrollments = class_instance.enrollment_set.filter(
            student__batch=expected_batch
        ).select_related('student')
        
        # Mapping student ID to enrollment object
        enrollment_map = {str(e.student.student_id): e for e in enrollments}
        
        updated_count = 0
        errors = []

        # Skip header
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[id_idx] is None:
                continue
                
            student_id = str(row[id_idx]).strip()
            score_val = row[score_idx]
            
            if not student_id or score_val is None:
                continue

            try:
                score_float = float(score_val)
            except (ValueError, TypeError):
                errors.append(f"Row {row_idx}: Student {student_id} — Invalid score value '{score_val}'")
                continue
            
            if score_float < 0:
                errors.append(f"Row {row_idx}: Student {student_id} — Score cannot be negative ({score_float})")
                continue
            
            if score_float > float(assessment.max_score):
                errors.append(f"Row {row_idx}: Student {student_id} — Score ({score_float}) exceeds max score ({assessment.max_score})")
                continue

            if student_id in enrollment_map:
                enrollment = enrollment_map[student_id]
                obj, created = StudentScore.objects.update_or_create(
                    enrollment=enrollment,
                    assessment=assessment,
                    defaults={'score': score_float}
                )
                updated_count += 1
            else:
                errors.append(f"Row {row_idx}: Student ID '{student_id}' is not enrolled in this class")
        
        if errors:
            error_list = "".join([f"<li>{err}</li>" for err in errors])
            messages.error(
                request, 
                f"Successfully imported scores for {updated_count} students. (Skipped/Invalid: {len(errors)})<br><ul>{error_list}</ul>",
                extra_tags='safe'
            )
        else:
            messages.success(request, f'Successfully imported scores for {updated_count} students.')
    
    except Exception as e:
        messages.error(request, f'Error reading Excel file: {str(e)}')
    
    return redirect('teacher_assignment_scores', class_id=class_id, assessment_id=assessment_id)


from django.contrib.auth.decorators import user_passes_test
from django.db import IntegrityError
from django.db.models import Q
from django.http import JsonResponse
from myapp.models import User, Department, Term, Subject, Enrollment

def is_frontdesk(user):
    return user.is_active and user.is_staff

@user_passes_test(is_frontdesk, login_url='/login')
def frontdesk_dashboard(request):
    total_students = Student.objects.count()
    total_classes = Class.objects.count()
    current_term = Term.objects.order_by('-start_date').first()
    total_departments = Department.objects.count()
    recent_students = Student.objects.select_related('user', 'department').order_by('-created_at')[:10]

    context = {
        'total_students': total_students,
        'total_classes': total_classes,
        'current_term': current_term,
        'total_departments': total_departments,
        'recent_students': recent_students,
    }
    return render(request, 'myapp/frontdesk_dashboard.html', context)

@user_passes_test(is_frontdesk, login_url='/login')
def frontdesk_register_student(request):
    departments = Department.objects.all().order_by('name')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        
        student_id = request.POST.get('student_id')
        department_id = request.POST.get('department')
        level = request.POST.get('level')
        batch = request.POST.get('batch')
        
        if not all([username, email, password, student_id, department_id, level]):
            messages.error(request, 'Please fill in all required fields.')
            return redirect('frontdesk_register_student')
            
        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" is already taken.')
            return redirect('frontdesk_register_student')
            
        if Student.objects.filter(student_id=student_id).exists():
            messages.error(request, f'Student ID "{student_id}" is already in use.')
            return redirect('frontdesk_register_student')
            
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            department = Department.objects.get(id=department_id)
            student = Student.objects.create(
                user=user,
                student_id=student_id,
                department=department,
                level=level,
                batch=batch
            )
            messages.success(request, f'Student {first_name} {last_name} ({student_id}) registered successfully!')
        except Exception as e:
            messages.error(request, f'Error registering student: {e}')
            
        return redirect('frontdesk_register_student')

    next_student_id = "ST-0001"
    latest_batch = "Batch 1"
    
    latest_student = Student.objects.order_by('-id').first()
    if latest_student:
        import re
        match = re.search(r'\d+', latest_student.student_id)
        if match:
            num = int(match.group())
            next_student_id = f"ST-{num + 1:04d}"
        
        if latest_student.batch:
            latest_batch = latest_student.batch

    context = {
        'departments': departments,
        'next_student_id': next_student_id,
        'latest_batch': latest_batch,
    }
    return render(request, 'myapp/frontdesk_register_student.html', context)

@user_passes_test(is_frontdesk, login_url='/login')
def frontdesk_enroll_student(request):
    prefill_student_id = request.GET.get('student_id', '')
    terms = Term.objects.order_by('-start_date')
    students = Student.objects.select_related('user').order_by('student_id')
    
    if request.method == 'POST':
        student_id = request.POST.get('student')
        class_id = request.POST.get('class_instance')
        
        if student_id and class_id:
            try:
                student = Student.objects.get(id=student_id)
                class_instance = Class.objects.get(id=class_id)
                
                Enrollment.objects.create(
                    student=student,
                    class_instance=class_instance
                )
                messages.success(request, f'Successfully enrolled {student.user.get_full_name()} into {class_instance.subject.name}.')
            except IntegrityError:
                messages.error(request, 'This student is already enrolled in this class.')
            except (Student.DoesNotExist, Class.DoesNotExist):
                messages.error(request, 'Invalid student or class selected.')
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')
                
        return redirect('frontdesk_enroll_student')
        
    context = {
        'terms': terms,
        'students': students,
        'prefill_student_id': int(prefill_student_id) if prefill_student_id.isdigit() else None
    }
    return render(request, 'myapp/frontdesk_enroll_student.html', context)

@user_passes_test(is_frontdesk, login_url='/login')
def frontdesk_api_classes(request):
    term_id = request.GET.get('term_id')
    if term_id:
        classes = Class.objects.filter(term_id=term_id).select_related('subject', 'teacher__user')
        data = []
        for c in classes:
            schedule = c.schedule if c.schedule else "TBD"
            name = f"{c.subject.name} - {c.teacher.user.get_full_name()} ({schedule})"
            data.append({'id': c.id, 'name': name})
        return JsonResponse(data, safe=False)
    return JsonResponse([], safe=False)

@user_passes_test(is_frontdesk, login_url='/login')
def frontdesk_manage_students(request):
    query = request.GET.get('q', '')
    department_id = request.GET.get('department', '')
    level = request.GET.get('level', '')
    batch = request.GET.get('batch', '')

    students = Student.objects.select_related('user', 'department')
    
    if query:
        students = students.filter(
            Q(student_id__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query)
        )
        
    if department_id:
        students = students.filter(department_id=department_id)
        
    if level:
        students = students.filter(level=level)
        
    if batch:
        students = students.filter(batch=batch)
        
    students = students.order_by('-created_at')
    
    # Options for filter dropdowns
    departments = Department.objects.all().order_by('name')
    levels = Student.objects.values_list('level', flat=True).distinct().order_by('level')
    batches = Student.objects.exclude(batch__isnull=True).exclude(batch='').values_list('batch', flat=True).distinct().order_by('batch')

    from django.core.paginator import Paginator
    paginator = Paginator(students, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
        
    context = {
        'students': page_obj, 
        'page_obj': page_obj,
        'query': query,
        'department_id': department_id,
        'level': level,
        'batch': batch,
        'departments': departments,
        'levels': levels,
        'batches': batches,
    }
    return render(request, 'myapp/frontdesk_manage_students.html', context)
